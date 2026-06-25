#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import importlib.util
import io
import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Sequence


EVALUATED_REQUIREMENT_CATEGORY = "功能"
DEFAULT_SHARD_SIZE = 2
REQUIRED_SHEET_NAME = "Sheet1"
STATE_FILENAME = "run_state.json"
SKILL_DIR = Path(__file__).resolve().parents[1]
REFERENCES_DIR = SKILL_DIR / "references"
SCORING_GUIDE_FILE = REFERENCES_DIR / "scoring-guide.md"
TSV_SCHEMA_FILE = REFERENCES_DIR / "output-tsv-schema.md"
CATEGORY_FIELD_CANDIDATES = (
    "分类类型",
    "需求分类",
    "OR需求分类",
    "需求类型",
    "需求类别",
    "OR需求类型",
    "需求分类类型",
)
EXCLUDED_SOURCE_FIELDS = {
    "TDR需求编号",
    "TDR需求名称*",
    "TDR需求描述*",
    "TDS需求编号",
    "TDS需求名称*",
    "TDS需求描述*",
}

DEFAULT_DIMENSIONS = [
    {
        "key": "or_user_language",
        "name": "OR-用户语言描述",
        "weight": 14,
        "description": "需求描述是否采用用户语言，避免技术术语，让非技术人员也能理解",
    },
    {
        "key": "or_scenario",
        "name": "OR-应用场景",
        "weight": 14,
        "description": "是否清晰描述了用户需求的应用场景，包括使用环境和上下文",
    },
    {
        "key": "or_user_value",
        "name": "OR-用户价值",
        "weight": 12,
        "description": "是否讲清楚解决的用户问题和带来的用户价值",
    },
    {
        "key": "dr_security",
        "name": "DR-安全分析",
        "weight": 6,
        "description": "是否识别涉及安全红线的业务部分，并说明其需要满足的安全要求",
    },
    {
        "key": "dr_testability",
        "name": "DR-可测试性",
        "weight": 16,
        "description": "是否使用技术语言清楚描述行为、输入、输出、条件、边界和结果，能够直接导出测试与验收标准",
    },
    {
        "key": "dr_ambiguity",
        "name": "DR-无歧义性",
        "weight": 10,
        "description": "术语、对象、范围、状态和结果判定是否只有一种合理理解",
    },
    {
        "key": "dr_exception",
        "name": "DR-异常描述",
        "weight": 8,
        "description": "是否明确异常路径、错误条件、非法输入处理、失败行为和边界场景",
    },
    {
        "key": "cross_scope",
        "name": "需求分解完整性",
        "weight": 7,
        "description": "OR 显式提出的关键目标、子目标和关键场景是否都被 DR 正面承接，是否存在明显漏拆",
    },
    {
        "key": "cross_dependencies",
        "name": "需求分解边界清晰度",
        "weight": 6,
        "description": "各 DR 的职责边界是否清楚，核心承接 DR 与支撑型 DR 是否区分明确，是否存在交叉或混写",
    },
    {
        "key": "cross_traceability",
        "name": "需求映射一致性",
        "weight": 7,
        "description": "各 DR 是否在语义上直接回应 OR 的目标、对象、条件和结果，而不是只有主题相关",
    },
]
DIMENSION_COLUMN_SPECS = tuple(
    {
        "key": str(item["key"]),
        "name": str(item["name"]),
        "weight": int(item["weight"]),
    }
    for item in DEFAULT_DIMENSIONS
)
TSV_COLUMNS = [
    "or_id",
    "or_name",
    "total_score",
    "or_score",
    "dr_average_score",
    "traceability_score",
    *[f"{item['key']}_score" for item in DIMENSION_COLUMN_SPECS],
    "grade",
    "weak_dimensions",
    "red_flags",
    "missing_items",
    "revision_actions",
    "evidence_summary",
]

CORE_FIELD_MAP = {
    "or_id": "OR需求编号",
    "or_name": "OR需求名称*",
    "or_desc": "OR需求描述*",
    "scenario": "应用场景",
    "customer_problem": "客户问题",
    "value_desc": "价值描述",
    "constraints": "约束与限制",
    "requirement_source": "需求来源",
    "region": "国家/区域",
    "dr_id": "DR需求编号",
    "dr_name": "DR需求名称*",
    "dr_desc": "DR需求描述*",
    "dr_integration": "集成方式",
    "dr_param": "参数规格",
    "dr_operation": "操作场景",
    "dr_test": "系统测试要点",
    "dr_security": "安全约束",
    "ds_id": "DS需求编号",
    "ds_name": "DS需求名称*",
    "ds_desc": "DS需求描述*",
    "spec_type": "规格分类",
    "subsystem": "所属子系统",
    "ds_dependencies": "假设和依赖信息",
    "ds_verify": "验证方法描述",
}

OR_CORE_ALIASES = (
    "or_id",
    "or_name",
    "or_desc",
    "scenario",
    "customer_problem",
    "value_desc",
    "constraints",
    "requirement_source",
    "region",
    "requirement_category",
)

DR_CORE_ALIASES = (
    "dr_id",
    "dr_name",
    "dr_desc",
    "dr_integration",
    "dr_param",
    "dr_operation",
    "dr_test",
    "dr_security",
    "ds_id",
    "ds_name",
    "ds_desc",
    "spec_type",
    "subsystem",
    "ds_dependencies",
    "ds_verify",
)

DIMENSION_FIELD_MAP = {
    "or_user_language": ["OR需求名称*", "OR需求描述*", "更多描述信息"],
    "or_scenario": ["应用场景", "操作场景", "OR需求描述*", "更多描述信息"],
    "or_user_value": ["客户问题", "价值描述", "需求来源", "OR需求描述*"],
    "dr_security": ["安全约束", "DR需求描述*", "更多描述信息"],
    "dr_testability": ["系统测试要点", "验证方法描述", "参数规格", "DR需求描述*", "集成方式"],
    "dr_ambiguity": ["参数规格", "DR需求描述*", "DS需求描述*"],
    "dr_exception": ["DR需求描述*", "系统测试要点", "验证方法描述", "安全约束", "更多描述信息"],
    "cross_scope": ["OR需求描述*", "DR需求名称*", "DR需求描述*"],
    "cross_dependencies": ["DR需求名称*", "DR需求描述*", "所属子系统", "假设和依赖信息", "集成方式", "规格分类"],
    "cross_traceability": ["OR需求编号", "OR需求名称*", "OR需求描述*", "DR需求编号", "DR需求名称*", "DR需求描述*", "ORURL", "DRURL"],
}


@dataclass
class RowRecord:
    index: int
    grouped: Dict[str, List[str]]

    def first(self, key: str, occurrence: int = 0) -> str:
        values = self.grouped.get(key, [])
        if occurrence < len(values):
            return clean_text(values[occurrence])
        return ""


@dataclass
class ReadResult:
    records: List[RowRecord]
    source_info: Dict[str, object]


def clean_text(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip())


def build_dimensions() -> List[Dict[str, object]]:
    return [dict(item) for item in DEFAULT_DIMENSIONS]


def dimension_column_name(key: str) -> str:
    return f"{key}_score"


def dimension_score_spec_by_key() -> Dict[str, Dict[str, object]]:
    return {str(item["key"]): dict(item) for item in DIMENSION_COLUMN_SPECS}


def dimension_score_spec_by_column() -> Dict[str, Dict[str, object]]:
    return {
        dimension_column_name(str(item["key"])): dict(item)
        for item in DIMENSION_COLUMN_SPECS
    }


def sanitize_grouped(grouped: Dict[str, List[str]]) -> Dict[str, List[str]]:
    return {
        key: values
        for key, values in grouped.items()
        if key not in EXCLUDED_SOURCE_FIELDS
    }


def missing_runtime_dependencies(path: Path) -> List[str]:
    suffix = path.suffix.lower()
    missing = []
    if suffix in {".xlsx", ".xlsm"} and importlib.util.find_spec("openpyxl") is None:
        missing.append("openpyxl")
    return missing


def dependency_install_hint(packages: Sequence[str]) -> str:
    joined = " ".join(packages)
    return f"python3 -m pip install {joined}"


def ensure_runtime_dependencies(path: Path) -> None:
    missing = missing_runtime_dependencies(path)
    if not missing:
        return
    package_list = ", ".join(missing)
    hint = dependency_install_hint(missing)
    raise SystemExit(f"缺少运行依赖: {package_list}。请先执行: {hint}")


def read_excel(path: Path) -> ReadResult:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("读取 Excel 需要安装 openpyxl") from exc
    workbook = load_workbook(path, read_only=False, data_only=True)
    if REQUIRED_SHEET_NAME not in workbook.sheetnames:
        raise SystemExit(f"Excel 输入必须包含工作表 {REQUIRED_SHEET_NAME}")
    sheet = workbook[REQUIRED_SHEET_NAME]
    source_info = {
        "input_format": path.suffix.lower().lstrip("."),
        "sheet_name": sheet.title,
    }
    if sheet.max_row < 1:
        return ReadResult(records=[], source_info=source_info)
    merged_values = {}
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        anchor_value = sheet.cell(min_row, min_col).value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_values[(row, col)] = anchor_value

    def cell_value(row: int, col: int) -> str:
        value = sheet.cell(row, col).value
        if value is None and (row, col) in merged_values:
            value = merged_values[(row, col)]
        return "" if value is None else str(value)

    header = [cell_value(1, col) for col in range(1, sheet.max_column + 1)]
    while header and not clean_text(header[-1]):
        header.pop()
    if not header:
        return ReadResult(records=[], source_info=source_info)

    records = []
    for row_index in range(2, sheet.max_row + 1):
        grouped: Dict[str, List[str]] = defaultdict(list)
        values = [cell_value(row_index, col) for col in range(1, len(header) + 1)]
        for name, value in zip(header, values):
            grouped[name].append(value)
        records.append(RowRecord(index=row_index - 1, grouped=sanitize_grouped(dict(grouped))))
    return ReadResult(records=records, source_info=source_info)


def read_json(path: Path) -> ReadResult:
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, dict):
        for value in data.values():
            if isinstance(value, list):
                data = value
                break
    if not isinstance(data, list):
        raise SystemExit("JSON 输入必须是对象数组，或包含数组字段的对象")
    records = []
    for idx, item in enumerate(data, start=1):
        if not isinstance(item, dict):
            continue
        grouped = sanitize_grouped(
            {str(key): ["" if value is None else str(value)] for key, value in item.items()}
        )
        records.append(RowRecord(index=idx, grouped=grouped))
    return ReadResult(
        records=records,
        source_info={
            "input_format": path.suffix.lower().lstrip("."),
        },
    )


def read_records(path: Path) -> ReadResult:
    ensure_runtime_dependencies(path)
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return read_excel(path)
    if suffix == ".json":
        return read_json(path)
    raise SystemExit(f"不支持的输入格式: {suffix}")


def summarize_headers(records: Sequence[RowRecord]) -> List[str]:
    if not records:
        return []
    headers = []
    seen = set()
    for record in records:
        for key in record.grouped:
            if key in EXCLUDED_SOURCE_FIELDS:
                continue
            if key not in seen:
                seen.add(key)
                headers.append(key)
    return headers


def extract_core_fields(record: RowRecord) -> Dict[str, str]:
    core = {}
    for alias, source in CORE_FIELD_MAP.items():
        core[alias] = record.first(source)
    core["requirement_category"] = requirement_category(record)["value"]
    return core


def select_core_fields(core_fields: Dict[str, str], aliases: Sequence[str]) -> Dict[str, str]:
    return {alias: core_fields.get(alias, "") for alias in aliases}


def filter_dimensions(dimensions: Sequence[Dict[str, object]], prefix: str) -> List[Dict[str, object]]:
    return [dimension for dimension in dimensions if str(dimension["key"]).startswith(prefix)]


def score_structure(dimensions: Sequence[Dict[str, object]]) -> Dict[str, int]:
    return {
        "or_total_weight": sum(int(d["weight"]) for d in dimensions if str(d["key"]).startswith("or_")),
        "dr_total_weight": sum(int(d["weight"]) for d in dimensions if str(d["key"]).startswith("dr_")),
        "cross_total_weight": sum(int(d["weight"]) for d in dimensions if str(d["key"]).startswith("cross_")),
    }


def merge_records(records: Sequence[RowRecord]) -> RowRecord:
    grouped: Dict[str, List[str]] = defaultdict(list)
    for record in records:
        for key, values in record.grouped.items():
            if key in EXCLUDED_SOURCE_FIELDS:
                continue
            for value in values:
                text = clean_text(value)
                if not text:
                    continue
                if text not in grouped[key]:
                    grouped[key].append(text)
    return RowRecord(index=records[0].index if records else 0, grouped=dict(grouped))


def build_raw_fields(record: RowRecord) -> Dict[str, List[str]]:
    return {
        key: [clean_text(value) for value in values if clean_text(value)]
        for key, values in record.grouped.items()
        if key not in EXCLUDED_SOURCE_FIELDS
        if any(clean_text(value) for value in values)
    }


def requirement_category(record: RowRecord) -> Dict[str, str]:
    for field in CATEGORY_FIELD_CANDIDATES:
        values = [clean_text(value) for value in record.grouped.get(field, []) if clean_text(value)]
        if values:
            return {"field": field, "value": values[0]}
    return {"field": "", "value": ""}


def category_display_name(category: str) -> str:
    return category or "未分类"


def is_evaluated_category(category: str) -> bool:
    return clean_text(category) == EVALUATED_REQUIREMENT_CATEGORY


def build_category_counts(or_groups: Sequence[Sequence[RowRecord]]) -> List[Dict[str, object]]:
    total = len(or_groups)
    counts: Dict[str, int] = {}
    fields: Dict[str, str] = {}

    for or_records in or_groups:
        merged_record = merge_records(or_records)
        category = requirement_category(merged_record)
        value = category_display_name(category["value"])
        counts[value] = counts.get(value, 0) + 1
        if category["field"] and value not in fields:
            fields[value] = category["field"]

    category_counts = []
    for value, count in counts.items():
        included = is_evaluated_category(value)
        percentage = round((count / total * 100), 2) if total else 0
        category_counts.append(
            {
                "category": value,
                "count": count,
                "percentage": percentage,
                "included_in_evaluation": included,
                "exclusion_reason": "" if included else f"分类类型不是{EVALUATED_REQUIREMENT_CATEGORY}",
                "source_field": fields.get(value, ""),
            }
        )
    return category_counts


def build_or_review_skeleton(score_weights: Dict[str, int], dr_items: Sequence[Dict[str, object]]) -> Dict[str, object]:
    return {
        "or_total_score": {
            "max_score": score_weights["or_total_weight"] + score_weights["dr_total_weight"] + score_weights["cross_total_weight"],
            "score": None,
            "review_conclusion": None,
        },
        "or_part": {
            "max_score": score_weights["or_total_weight"],
            "score": None,
            "dimension_scores": [],
        },
        "dr_parts": [
            {
                "dr_id": item["id"],
                "dr_name": item["name"],
                "max_score": score_weights["dr_total_weight"],
                "score": None,
                "dimension_scores": [],
            }
            for item in dr_items
        ],
        "dr_average": {
            "max_score": score_weights["dr_total_weight"],
            "score": None,
        },
        "decomposition_quality": {
            "max_score": score_weights["cross_total_weight"],
            "score": None,
            "dimension_scores": [],
        },
        "review_decision": {
            "blocking_issues": [],
            "triggered_red_line_rules": [],
        },
    }


def group_records_by_or(records: Sequence[RowRecord]) -> List[List[RowRecord]]:
    groups: List[List[RowRecord]] = []
    current: List[RowRecord] = []
    current_key = ""

    for record in records:
        core = extract_core_fields(record)
        or_key = core.get("or_id") or ""
        if current and or_key and or_key != current_key:
            groups.append(current)
            current = [record]
            current_key = or_key
            continue
        if current:
            current.append(record)
            if or_key:
                current_key = or_key
            continue
        current = [record]
        current_key = or_key or f"ROW-{record.index}"

    if current:
        groups.append(current)
    return groups


def group_records_by_dr(records: Sequence[RowRecord]) -> List[List[RowRecord]]:
    groups: List[List[RowRecord]] = []
    current: List[RowRecord] = []
    current_key = ""

    for record in records:
        core = extract_core_fields(record)
        dr_key = core.get("dr_id") or f"ROW-{record.index}"
        if current and dr_key != current_key:
            groups.append(current)
            current = [record]
            current_key = dr_key
            continue
        if current:
            current.append(record)
            continue
        current = [record]
        current_key = dr_key

    if current:
        groups.append(current)
    return groups


def build_dimension_view(record: RowRecord, dimensions: Sequence[Dict[str, object]]) -> Dict[str, Dict[str, object]]:
    view = {}
    for dimension in dimensions:
        key = str(dimension["key"])
        mapped_fields = DIMENSION_FIELD_MAP.get(key, [])
        evidence_fields = {}
        for field in mapped_fields:
            values = [clean_text(value) for value in record.grouped.get(field, []) if clean_text(value)]
            if values:
                evidence_fields[field] = values
        view[key] = {
            "name": dimension["name"],
            "mapped_fields": mapped_fields,
            "evidence_fields": evidence_fields,
        }
    return view


def build_review_packet(
    input_path: Path,
    dimensions: List[Dict[str, object]],
    records: Sequence[RowRecord],
    source_info: Dict[str, object] | None = None,
) -> Dict[str, object]:
    or_dimensions = filter_dimensions(dimensions, "or_")
    dr_dimensions = filter_dimensions(dimensions, "dr_")
    cross_dimensions = filter_dimensions(dimensions, "cross_")

    score_weights = score_structure(dimensions)
    or_record_groups = group_records_by_or(records)
    all_category_counts = build_category_counts(or_record_groups)
    groups = []
    total_dr_count = 0
    for or_records in or_record_groups:
        merged_or_record = merge_records(or_records)
        raw_fields = build_raw_fields(merged_or_record)
        if not raw_fields:
            continue
        category = requirement_category(merged_or_record)
        if not is_evaluated_category(category["value"]):
            continue
        full_core_fields = extract_core_fields(merged_or_record)
        or_core_fields = select_core_fields(full_core_fields, OR_CORE_ALIASES)
        requirement_id = full_core_fields.get("or_id") or full_core_fields.get("dr_id") or full_core_fields.get("ds_id") or f"ROW-{or_records[0].index}"
        requirement_name = full_core_fields.get("or_name") or full_core_fields.get("dr_name") or requirement_id

        dr_items = []
        for dr_records in group_records_by_dr(or_records):
            merged_dr_record = merge_records(dr_records)
            full_dr_core_fields = extract_core_fields(merged_dr_record)
            dr_core_fields = select_core_fields(full_dr_core_fields, DR_CORE_ALIASES)
            dr_id = full_dr_core_fields.get("dr_id") or f"ROW-{dr_records[0].index}"
            dr_name = full_dr_core_fields.get("dr_name") or dr_id
            dr_raw_fields = build_raw_fields(merged_dr_record)
            if not dr_raw_fields:
                continue
            dr_items.append(
                {
                    "row_indices": [record.index for record in dr_records],
                    "id": dr_id,
                    "name": dr_name,
                    "core_fields": dr_core_fields,
                    "dimension_view": build_dimension_view(merged_dr_record, dr_dimensions),
                    "raw_fields": dr_raw_fields,
                }
            )
        total_dr_count += len(dr_items)

        groups.append(
            {
                "row_indices": [record.index for record in or_records],
                "id": requirement_id,
                "name": requirement_name,
                "category": category_display_name(category["value"]),
                "category_field": category["field"],
                "or_core_fields": or_core_fields,
                "or_dimension_view": build_dimension_view(merged_or_record, or_dimensions),
                "dr_items": dr_items,
                "dr_count": len(dr_items),
                "cross_dimension_view": build_dimension_view(merged_or_record, cross_dimensions),
                "review_skeleton": build_or_review_skeleton(score_weights, dr_items),
                "raw_fields": raw_fields,
            }
        )

    return {
        "input_path": str(input_path),
        "source_info": dict(source_info or {}),
        "score_structure": score_weights,
        "evaluation_filter": {
            "category_field_candidates": list(CATEGORY_FIELD_CANDIDATES),
            "included_category": EVALUATED_REQUIREMENT_CATEGORY,
            "rule": f"only OR units whose requirement category is exactly `{EVALUATED_REQUIREMENT_CATEGORY}` are included in evaluation",
        },
        "all_category_counts": all_category_counts,
        "excluded_or_count": sum(
            int(item["count"]) for item in all_category_counts if not bool(item["included_in_evaluation"])
        ),
        "item_count": len(groups),
        "or_count": len(groups),
        "dr_count": total_dr_count,
        "dimension_count": len(dimensions),
        "dimensions": dimensions,
        "header_summary": summarize_headers(records),
        "groups": groups,
    }


def read_json_file(path: Path) -> Dict[str, object]:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json_file(path: Path, payload: Dict[str, object]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def default_state_path(workspace_dir: Path) -> Path:
    return workspace_dir / STATE_FILENAME


def build_run_state(
    *,
    workspace_dir: Path,
    packet_path: Path,
    shard_records: Sequence[Dict[str, Path]],
) -> Dict[str, object]:
    shards = []
    for record in shard_records:
        shard_path = Path(record["shard"]).resolve()
        prompt_path = Path(record["prompt"]).resolve()
        shard_id = shard_path.stem
        shards.append(
            {
                "shard_id": shard_id,
                "status": "pending",
                "next_action": "evaluate",
                "attempt_counts": {"raw": 0, "repair": 0, "rerun": 0},
                "paths": {
                    "shard": str(shard_path),
                    "prompt": str(prompt_path),
                    "raw_output": str((workspace_dir / "results" / f"{shard_id}.raw.tsv").resolve()),
                    "valid_output": str((workspace_dir / "results" / f"{shard_id}.valid.json").resolve()),
                    "repair_prompt": str((workspace_dir / "repairs" / f"{shard_id}.repair_prompt.txt").resolve()),
                },
                "last_attempt_kind": "",
                "last_errors": [],
            }
        )
    return {
        "workspace_dir": str(workspace_dir.resolve()),
        "packet_path": str(packet_path.resolve()),
        "overall_status": "pending",
        "shards": shards,
        "report_path": "",
        "summary_json_path": "",
    }


def load_run_state(path: Path) -> Dict[str, object]:
    return read_json_file(path)


def save_run_state(path: Path, state: Dict[str, object]) -> None:
    write_json_file(path, state)


def find_shard_state(state: Dict[str, object], shard_id: str) -> Dict[str, object]:
    for item in state.get("shards", []):
        if str(item.get("shard_id", "")) == shard_id:
            return item
    raise KeyError(f"shard_id not found in state: {shard_id}")


def recompute_overall_status(state: Dict[str, object]) -> None:
    shard_statuses = [str(item.get("status", "")) for item in state.get("shards", [])]
    if shard_statuses and all(status == "valid" for status in shard_statuses):
        state["overall_status"] = "ready_to_aggregate"
        return
    if any(status == "rerun_required" for status in shard_statuses):
        state["overall_status"] = "rerun_required"
        return
    if any(status == "repairable" for status in shard_statuses):
        state["overall_status"] = "repair_required"
        return
    if any(status == "valid" for status in shard_statuses):
        state["overall_status"] = "in_progress"
        return
    state["overall_status"] = "pending"


def update_run_state_for_validation(
    *,
    state: Dict[str, object],
    shard_id: str,
    validation: Dict[str, object],
    raw_output_path: Path,
    output_path: Path,
    repair_prompt_path: Path | None,
    attempt_kind: str,
) -> None:
    shard_state = find_shard_state(state, shard_id)
    attempt_counts = shard_state.setdefault("attempt_counts", {"raw": 0, "repair": 0, "rerun": 0})
    attempt_counts[attempt_kind] = int(attempt_counts.get(attempt_kind, 0)) + 1
    shard_state["status"] = str(validation.get("status", ""))
    shard_state["last_attempt_kind"] = attempt_kind
    shard_state["last_errors"] = list(validation.get("errors", []))
    paths = shard_state.setdefault("paths", {})
    paths["last_raw_output"] = str(raw_output_path.resolve())
    paths["last_validation_output"] = str(output_path.resolve())
    if repair_prompt_path is not None:
        paths["repair_prompt"] = str(repair_prompt_path.resolve())

    if shard_state["status"] == "valid":
        shard_state["next_action"] = "none"
    elif shard_state["status"] == "repairable":
        shard_state["next_action"] = "repair"
    elif shard_state["status"] == "rerun_required":
        shard_state["next_action"] = "rerun"
    else:
        shard_state["next_action"] = "inspect"
    recompute_overall_status(state)


def resolve_state_path_from_artifact(path: Path) -> Path | None:
    candidate = path.parents[1] / STATE_FILENAME
    if candidate.exists():
        return candidate.resolve()
    return None


def build_packet_from_input(input_path: Path) -> Dict[str, object]:
    dimensions = build_dimensions()
    read_result = read_records(input_path)
    return build_review_packet(
        input_path,
        dimensions,
        read_result.records,
        source_info=read_result.source_info,
    )


def build_shards(packet: Dict[str, object], shard_size: int = DEFAULT_SHARD_SIZE) -> List[Dict[str, object]]:
    if shard_size < 1:
        raise ValueError("shard_size must be at least 1")
    groups = list(packet.get("groups", []))
    shards = []
    shard_count = (len(groups) + shard_size - 1) // shard_size
    for index in range(shard_count):
        chunk = groups[index * shard_size : (index + 1) * shard_size]
        shard_id = f"shard_{index + 1:04d}"
        shards.append(
            {
                "shard_id": shard_id,
                "shard_index": index + 1,
                "shard_count": shard_count,
                "expected_or_ids": [str(item.get("id", "")) for item in chunk],
                "input_path": packet.get("input_path", ""),
                "source_info": packet.get("source_info", {}),
                "score_structure": packet.get("score_structure", {}),
                "dimensions": packet.get("dimensions", []),
                "groups": chunk,
            }
        )
    return shards


def render_shard_prompt(shard: Dict[str, object]) -> str:
    scoring_guide = SCORING_GUIDE_FILE.read_text(encoding="utf-8")
    output_schema = TSV_SCHEMA_FILE.read_text(encoding="utf-8")
    shard_json = json.dumps(shard, ensure_ascii=False, indent=2)
    return (
        "You are evaluating one independent requirements shard.\n"
        "Use only the evidence in SHARD_JSON. Do not use prior conversation context.\n"
        "Apply the complete SCORING_GUIDE and output exactly the TSV required by OUTPUT_TSV_SCHEMA.\n"
        "Do not treat a blank or omitted spreadsheet column as a requirement gap by itself.\n"
        "Judge missing_items from the substantive OR/DR/DS requirement text and other non-empty evidence in SHARD_JSON.\n"
        "If a dedicated Excel column is blank but the same meaning is stated elsewhere in the requirement description, do not mark it as missing for that reason.\n"
        "Do not output Markdown, explanations, JSON, code fences, or any text outside the TSV.\n\n"
        "[SCORING_GUIDE]\n"
        f"{scoring_guide}\n\n"
        "[OUTPUT_TSV_SCHEMA]\n"
        f"{output_schema}\n\n"
        "[SHARD_JSON]\n"
        f"{shard_json}\n"
    )


def render_repair_prompt(raw_output: str) -> str:
    output_schema = TSV_SCHEMA_FILE.read_text(encoding="utf-8")
    return (
        "Your previous output is not valid TSV.\n"
        "Convert it to the exact TSV schema below.\n"
        "Do not change scores, conclusions, or wording meaning.\n"
        "Do not re-evaluate the requirements.\n"
        "Only output TSV. Do not output Markdown, explanations, JSON, or code fences.\n\n"
        "[OUTPUT_TSV_SCHEMA]\n"
        f"{output_schema}\n\n"
        "[INVALID_OUTPUT]\n"
        f"{raw_output.strip()}\n"
    )


def write_shards_and_prompts(packet: Dict[str, object], out_dir: Path, shard_size: int) -> List[Dict[str, Path]]:
    shards_dir = out_dir / "shards"
    prompts_dir = out_dir / "prompts"
    shards_dir.mkdir(parents=True, exist_ok=True)
    prompts_dir.mkdir(parents=True, exist_ok=True)

    records = []
    for shard in build_shards(packet, shard_size):
        shard_id = str(shard["shard_id"])
        shard_path = shards_dir / f"{shard_id}.json"
        prompt_path = prompts_dir / f"{shard_id}.prompt.txt"
        write_json_file(shard_path, shard)
        prompt_path.write_text(render_shard_prompt(shard), encoding="utf-8")
        records.append({"shard": shard_path, "prompt": prompt_path})
    return records


def parse_number(value: str, field: str, minimum: float, maximum: float, errors: List[str]) -> float | None:
    try:
        number = float(str(value).strip())
    except ValueError:
        errors.append(f"{field} is not a number: {value}")
        return None
    if number < minimum or number > maximum:
        errors.append(f"{field} out of range {minimum}-{maximum}: {value}")
    return number


def grade_for_score(score: float) -> str:
    if score >= 90:
        return "excellent"
    if score >= 75:
        return "good"
    if score >= 60:
        return "fair"
    return "poor"


def compact_number(value: float) -> int | float:
    if value.is_integer():
        return int(value)
    return round(value, 2)


def normalized_dimension_score(score: float, max_score: int) -> float:
    if max_score <= 0:
        return 0.0
    return round(score / max_score * 100, 2)


def split_list_field(value: str) -> List[str]:
    text = clean_text(value)
    if text in {"", "无", "none", "None", "N/A", "n/a"}:
        return []
    return [item.strip() for item in text.split(";") if item.strip()]


def strict_parse_tsv(raw_output: str) -> List[Dict[str, str]]:
    text = raw_output.strip()
    if not text:
        raise ValueError("output is empty")
    if text.startswith("```") or text.endswith("```"):
        raise ValueError("output must not use code fences")
    reader = csv.DictReader(io.StringIO(text), delimiter="\t")
    if reader.fieldnames != TSV_COLUMNS:
        raise ValueError("TSV header does not match the required schema")
    rows = list(reader)
    if not rows:
        raise ValueError("TSV contains no result rows")
    return rows


def looks_repairable(raw_output: str, expected_or_ids: Sequence[str]) -> bool:
    text = raw_output.strip()
    if not text:
        return False
    if not all(or_id and or_id in text for or_id in expected_or_ids):
        return False
    schema_signals = sum(1 for column in ("or_id", "total_score", "or_score", "grade") if column in text)
    format_signals = int("|" in text) + int("\t" in text) + int("```" in text)
    return schema_signals >= 2 or format_signals >= 1


def validate_tsv_output(shard: Dict[str, object], raw_output: str) -> Dict[str, object]:
    expected_or_ids = [str(item) for item in shard.get("expected_or_ids", [])]
    try:
        rows = strict_parse_tsv(raw_output)
    except ValueError as exc:
        status = "repairable" if looks_repairable(raw_output, expected_or_ids) else "rerun_required"
        return {
            "status": status,
            "shard_id": shard.get("shard_id", ""),
            "errors": [str(exc)],
            "results": [],
        }

    errors: List[str] = []
    row_ids = [clean_text(row.get("or_id", "")) for row in rows]
    if len(rows) != len(expected_or_ids):
        errors.append(f"expected {len(expected_or_ids)} OR rows, got {len(rows)}")
    if row_ids != expected_or_ids:
        errors.append(f"OR ids do not match shard order: expected {expected_or_ids}, got {row_ids}")

    results = []
    dimension_specs = dimension_score_spec_by_column()
    for row in rows:
        row_errors: List[str] = []
        total_score = parse_number(row.get("total_score", ""), "total_score", 0, 100, row_errors)
        or_score = parse_number(row.get("or_score", ""), "or_score", 0, 40, row_errors)
        dr_average_score = parse_number(row.get("dr_average_score", ""), "dr_average_score", 0, 40, row_errors)
        traceability_score = parse_number(row.get("traceability_score", ""), "traceability_score", 0, 20, row_errors)
        dimension_scores: Dict[str, int | float] = {}
        or_dimension_total = 0.0
        dr_dimension_total = 0.0
        cross_dimension_total = 0.0
        for column, spec in dimension_specs.items():
            score = parse_number(row.get(column, ""), column, 0, int(spec["weight"]), row_errors)
            if score is None:
                continue
            key = str(spec["key"])
            dimension_scores[key] = compact_number(float(score))
            if key.startswith("or_"):
                or_dimension_total += float(score)
            elif key.startswith("dr_"):
                dr_dimension_total += float(score)
            else:
                cross_dimension_total += float(score)
        grade = clean_text(row.get("grade", "")).lower()
        if grade not in {"excellent", "good", "fair", "poor"}:
            row_errors.append(f"invalid grade: {row.get('grade', '')}")
        if None not in {total_score, or_score, dr_average_score, traceability_score}:
            computed_total = float(or_score) + float(dr_average_score) + float(traceability_score)
            if abs(float(total_score) - computed_total) > 0.05:
                row_errors.append(
                    "total_score must equal or_score + dr_average_score + traceability_score"
                )
            if abs(float(or_score) - or_dimension_total) > 0.05:
                row_errors.append("or_score must equal the sum of OR dimension scores")
            if abs(float(dr_average_score) - dr_dimension_total) > 0.05:
                row_errors.append("dr_average_score must equal the sum of DR dimension scores")
            if abs(float(traceability_score) - cross_dimension_total) > 0.05:
                row_errors.append("traceability_score must equal the sum of cross dimension scores")
            expected_grade = grade_for_score(float(total_score))
            if grade and grade != expected_grade:
                row_errors.append(f"grade {grade} does not match total_score {total_score}; expected {expected_grade}")
        summary = clean_text(row.get("evidence_summary", ""))
        if len(summary) > 240:
            row_errors.append("evidence_summary is too long")
        if row_errors:
            errors.extend([f"{row.get('or_id', '')}: {message}" for message in row_errors])
            continue
        results.append(
            {
                "or_id": clean_text(row.get("or_id", "")),
                "or_name": clean_text(row.get("or_name", "")),
                "total_score": compact_number(float(total_score)),
                "or_score": compact_number(float(or_score)),
                "dr_average_score": compact_number(float(dr_average_score)),
                "traceability_score": compact_number(float(traceability_score)),
                "dimension_scores": dimension_scores,
                "grade": grade,
                "weak_dimensions": split_list_field(row.get("weak_dimensions", "")),
                "red_flags": split_list_field(row.get("red_flags", "")),
                "missing_items": split_list_field(row.get("missing_items", "")),
                "revision_actions": split_list_field(row.get("revision_actions", "")),
                "evidence_summary": summary,
            }
        )

    return {
        "status": "valid" if not errors else "rerun_required",
        "shard_id": shard.get("shard_id", ""),
        "expected_or_ids": expected_or_ids,
        "errors": errors,
        "results": results if not errors else [],
    }


def md_cell(value: object) -> str:
    return str(value).replace("\n", " ").replace("|", "\\|")


def count_values(results: Sequence[Dict[str, object]], field: str) -> List[tuple[str, int]]:
    counter: Counter[str] = Counter()
    for result in results:
        for value in result.get(field, []):
            if value:
                counter[str(value)] += 1
    return counter.most_common()


def load_valid_results(results_dir: Path) -> List[Dict[str, object]]:
    results = []
    for path in sorted(results_dir.glob("*.valid.json")):
        payload = read_json_file(path)
        if payload.get("status") != "valid":
            continue
        results.extend(payload.get("results", []))
    return results


def build_dimension_summary(results: Sequence[Dict[str, object]]) -> List[Dict[str, object]]:
    specs = dimension_score_spec_by_key()
    summary = []
    result_count = len(results)
    ordered_keys = [str(item["key"]) for item in DIMENSION_COLUMN_SPECS]
    for key in ordered_keys:
        spec = specs[key]
        scores = [
            float(result.get("dimension_scores", {}).get(key, 0))
            for result in results
            if key in result.get("dimension_scores", {})
        ]
        average_score = sum(scores) / len(scores) if scores else 0.0
        weak_count = sum(
            1
            for result in results
            if str(spec["name"]) in result.get("weak_dimensions", [])
        )
        weak_rate = weak_count / result_count if result_count else 0.0
        summary.append(
            {
                "key": key,
                "name": str(spec["name"]),
                "max_score": int(spec["weight"]),
                "average_score": compact_number(average_score),
                "normalized_average_score": normalized_dimension_score(average_score, int(spec["weight"])),
                "weak_count": weak_count,
                "weak_rate": round(weak_rate, 4),
            }
        )
    return summary


def top_dimension_items(
    dimension_summary: Sequence[Dict[str, object]],
    *,
    descending: bool,
    limit: int = 3,
) -> List[Dict[str, object]]:
    ranked = sorted(
        dimension_summary,
        key=lambda item: float(item["normalized_average_score"]),
        reverse=descending,
    )
    return ranked[:limit]


def build_report_summary(
    packet: Dict[str, object],
    results: Sequence[Dict[str, object]],
    report_path: Path,
) -> Dict[str, object]:
    groups = list(packet.get("groups", []))
    result_by_id = {str(item.get("or_id", "")): item for item in results}
    expected_ids = [str(item.get("id", "")) for item in groups]
    missing_ids = [or_id for or_id in expected_ids if or_id not in result_by_id]
    scored_results = [result_by_id[or_id] for or_id in expected_ids if or_id in result_by_id]
    average = (
        sum(float(item["total_score"]) for item in scored_results) / len(scored_results)
        if scored_results
        else 0
    )
    distribution = Counter(str(item.get("grade", "")) for item in scored_results)
    dimension_summary = build_dimension_summary(scored_results)
    input_path = str(packet.get("input_path", ""))
    return {
        "project_name": Path(input_path).stem if input_path else report_path.stem,
        "report_path": str(report_path),
        "input_path": input_path,
        "source_info": dict(packet.get("source_info", {})),
        "or_count": int(packet.get("or_count", 0)),
        "dr_count": int(packet.get("dr_count", 0)),
        "scored_or_count": len(scored_results),
        "unscored_or_count": len(missing_ids),
        "average_score": round(average, 2),
        "grade_distribution": {grade: distribution.get(grade, 0) for grade in ("excellent", "good", "fair", "poor")},
        "dimension_summary": dimension_summary,
        "strength_dimensions": top_dimension_items(dimension_summary, descending=True),
        "weak_dimensions": top_dimension_items(dimension_summary, descending=False),
        "weak_counts": count_values(scored_results, "weak_dimensions"),
        "missing_counts": count_values(scored_results, "missing_items"),
        "revision_counts": count_values(scored_results, "revision_actions"),
        "low_score_ors": [
            result
            for result in sorted(scored_results, key=lambda item: float(item["total_score"]))[:5]
        ],
        "missing_or_ids": missing_ids,
    }


def render_aggregate_report(packet: Dict[str, object], results: Sequence[Dict[str, object]]) -> str:
    report_summary = build_report_summary(packet, results, Path(str(packet.get("input_path", "report"))))
    groups = list(packet.get("groups", []))
    result_by_id = {str(item.get("or_id", "")): item for item in results}
    expected_ids = [str(item.get("id", "")) for item in groups]
    missing_ids = [or_id for or_id in expected_ids if or_id not in result_by_id]
    scored_results = [result_by_id[or_id] for or_id in expected_ids if or_id in result_by_id]
    average = (
        sum(float(item["total_score"]) for item in scored_results) / len(scored_results)
        if scored_results
        else 0
    )
    distribution = Counter(str(item.get("grade", "")) for item in scored_results)

    lines = [
        "# 需求评估报告",
        "",
        "## 1. 评估概览",
        "",
        f"- 数据源: `{packet.get('input_path', '')}`",
    ]
    for key, value in packet.get("source_info", {}).items():
        lines.append(f"- {key}: `{value}`")
        lines.extend(
        [
            f"- 参与评估 OR 数: {packet.get('or_count', 0)}",
            f"- 参与评估 DR 数: {packet.get('dr_count', 0)}",
            f"- 成功评估 OR 数: {len(scored_results)}",
            f"- 未评估 OR 数: {len(missing_ids)}",
            f"- 平均分: {average:.2f}",
            "- 评分结构: OR(用户语言描述 14 / 应用场景 14 / 用户价值 12) + DR(安全分析 6 / 可测试性 16 / 无歧义性 10 / 异常描述 8) + 分解映射(完整性 7 / 边界清晰度 6 / 映射一致性 7)",
            "",
            "## 2. OR需求分类统计",
            "",
            "| 需求分类 | OR条目数 | 占比 | 是否参与评审 | 排除原因 |",
            "| --- | ---: | ---: | --- | --- |",
        ]
    )
    for item in packet.get("all_category_counts", []):
        included = "是" if item.get("included_in_evaluation") else "否"
        lines.append(
            f"| {md_cell(item.get('category', ''))} | {item.get('count', 0)} | {item.get('percentage', 0)}% | {included} | {md_cell(item.get('exclusion_reason', '') or '')} |"
        )

    lines.extend(
        [
            "",
            "## 3. 分数分布",
            "",
            "| 等级 | 数量 |",
            "| --- | ---: |",
        ]
    )
    for grade in ("excellent", "good", "fair", "poor"):
        lines.append(f"| {grade} | {distribution.get(grade, 0)} |")

        lines.extend(
        [
            "",
            "## 4. 全 OR 分数表",
            "",
            "| OR编号 | OR名称 | 总分 | OR部分 | DR平均 | 分解映射 | 等级 |",
            "| --- | --- | ---: | ---: | ---: | ---: | --- |",
        ]
    )
    for group in groups:
        or_id = str(group.get("id", ""))
        result = result_by_id.get(or_id)
        if result is None:
            lines.append(f"| {md_cell(or_id)} | {md_cell(group.get('name', ''))} | 未评估 |  |  |  |  |")
            continue
        lines.append(
            f"| {md_cell(or_id)} | {md_cell(result.get('or_name', group.get('name', '')))} | {result.get('total_score', '')} | {result.get('or_score', '')} | {result.get('dr_average_score', '')} | {result.get('traceability_score', '')} | {result.get('grade', '')} |"
        )

    def append_counter_section(title: str, values: List[tuple[str, int]]) -> None:
        lines.extend(["", title, "", "| 项目 | 次数 |", "| --- | ---: |"])
        if not values:
            lines.append("| 无 | 0 |")
            return
        for value, count in values[:10]:
            lines.append(f"| {md_cell(value)} | {count} |")

    append_counter_section("## 5. 高频弱项", count_values(scored_results, "weak_dimensions"))
    lines.extend(
        [
            "",
            "## 6. 维度均分画像",
            "",
            "| 维度 | 平均得分 | 满分 | 标准化均分 | 弱项命中数 | 弱项命中率 |",
            "| --- | ---: | ---: | ---: | ---: | ---: |",
        ]
    )
    for item in report_summary["dimension_summary"]:
        lines.append(
            f"| {md_cell(item['name'])} | {item['average_score']} | {item['max_score']} | "
            f"{item['normalized_average_score']} | {item['weak_count']} | {item['weak_rate'] * 100:.2f}% |"
        )
    lines.extend(
        [
            "",
            f"- 优势维度: {'; '.join(str(item['name']) for item in report_summary['strength_dimensions']) or '无'}",
            f"- 薄弱维度: {'; '.join(str(item['name']) for item in report_summary['weak_dimensions']) or '无'}",
        ]
    )
    append_counter_section("## 7. 高频缺失项", count_values(scored_results, "missing_items"))
    append_counter_section("## 8. 高频修改建议", count_values(scored_results, "revision_actions"))

    lines.extend(["", "## 9. 低分 OR 详情", ""])
    for result in sorted(scored_results, key=lambda item: float(item["total_score"]))[:5]:
        dimension_score_text = "; ".join(
            f"{item['name']}={result.get('dimension_scores', {}).get(item['key'], 0)}/{item['max_score']}"
            for item in report_summary["dimension_summary"]
        )
        lines.extend(
            [
                f"### {result.get('or_id', '')} {result.get('or_name', '')}",
                "",
                f"- 总分: {result.get('total_score', '')}",
                f"- 等级: {result.get('grade', '')}",
                f"- 维度分: {dimension_score_text}",
                f"- 弱项: {'; '.join(result.get('weak_dimensions', [])) or '无'}",
                f"- 触发红线: {'; '.join(result.get('red_flags', [])) or '无'}",
                f"- 缺失项: {'; '.join(result.get('missing_items', [])) or '无'}",
                f"- 修改建议: {'; '.join(result.get('revision_actions', [])) or '无'}",
                f"- 证据摘要: {result.get('evidence_summary', '')}",
                "",
            ]
        )

    lines.extend(["## 10. 未评估或失败项", ""])
    if missing_ids:
        for or_id in missing_ids:
            lines.append(f"- {or_id}")
    else:
        lines.append("- 无")
    return "\n".join(lines) + "\n"


def command_build_packet(args: argparse.Namespace) -> None:
    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve() if args.output else input_path.with_suffix(".json")
    packet = build_packet_from_input(input_path)
    write_json_file(output_path, packet)
    print(f"已生成 packet: {output_path}")
    print(f"OR条目数: {packet['or_count']}")


def command_split_shards(args: argparse.Namespace) -> None:
    packet_path = Path(args.packet).expanduser().resolve()
    out_dir = Path(args.out_dir).expanduser().resolve()
    packet = read_json_file(packet_path)
    records = write_shards_and_prompts(packet, out_dir, args.shard_size)
    print(f"已生成 shards/prompts: {out_dir}")
    print(f"分片数: {len(records)}")


def command_prepare(args: argparse.Namespace) -> None:
    input_path = Path(args.input).expanduser().resolve()
    out_dir = Path(args.out_dir).expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    packet = build_packet_from_input(input_path)
    packet_path = out_dir / "packet.json"
    write_json_file(packet_path, packet)
    records = write_shards_and_prompts(packet, out_dir, args.shard_size)
    (out_dir / "results").mkdir(exist_ok=True)
    (out_dir / "repairs").mkdir(exist_ok=True)
    state_path = (
        Path(args.state_json).expanduser().resolve()
        if args.state_json
        else default_state_path(out_dir).resolve()
    )
    save_run_state(
        state_path,
        build_run_state(workspace_dir=out_dir, packet_path=packet_path, shard_records=records),
    )
    print(f"已生成 packet: {packet_path}")
    print(f"已生成分片数: {len(records)}")
    print(f"prompt目录: {out_dir / 'prompts'}")
    print(f"运行状态文件: {state_path}")


def command_validate_result(args: argparse.Namespace) -> None:
    shard_path = Path(args.shard).expanduser().resolve()
    raw_output_path = Path(args.raw_output).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()
    shard = read_json_file(shard_path)
    raw_output = raw_output_path.read_text(encoding="utf-8")
    validation = validate_tsv_output(shard, raw_output)
    write_json_file(output_path, validation)
    repair_prompt_path = None
    if validation["status"] == "repairable" and args.repair_prompt:
        repair_prompt_path = Path(args.repair_prompt).expanduser().resolve()
        repair_prompt_path.parent.mkdir(parents=True, exist_ok=True)
        repair_prompt_path.write_text(render_repair_prompt(raw_output), encoding="utf-8")
        print(f"已生成 repair prompt: {repair_prompt_path}")
    state_path = (
        Path(args.state_json).expanduser().resolve()
        if args.state_json
        else resolve_state_path_from_artifact(shard_path)
    )
    if state_path is not None and state_path.exists():
        state = load_run_state(state_path)
        update_run_state_for_validation(
            state=state,
            shard_id=str(shard.get("shard_id", shard_path.stem)),
            validation=validation,
            raw_output_path=raw_output_path,
            output_path=output_path,
            repair_prompt_path=repair_prompt_path,
            attempt_kind=args.attempt_kind,
        )
        save_run_state(state_path, state)
        print(f"已更新运行状态: {state_path}")
    print(f"validation_status: {validation['status']}")
    if validation.get("errors"):
        for error in validation["errors"]:
            print(f"- {error}")


def command_aggregate(args: argparse.Namespace) -> None:
    packet_path = Path(args.packet).expanduser().resolve()
    results_dir = Path(args.results_dir).expanduser().resolve()
    report_path = Path(args.report).expanduser().resolve()
    state_path = (
        Path(args.state_json).expanduser().resolve()
        if args.state_json
        else default_state_path(packet_path.parent).resolve()
    )
    if state_path.exists():
        state = load_run_state(state_path)
        non_valid = [
            str(item.get("shard_id", ""))
            for item in state.get("shards", [])
            if str(item.get("status", "")) != "valid"
        ]
        if non_valid:
            joined = ", ".join(non_valid)
            raise SystemExit(f"无法聚合，以下分片尚未通过验证: {joined}")
    packet = read_json_file(packet_path)
    results = load_valid_results(results_dir)
    report = render_aggregate_report(packet, results)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(report, encoding="utf-8")
    summary_path = (
        Path(args.summary_json).expanduser().resolve()
        if args.summary_json
        else report_path.with_suffix(".summary.json")
    )
    write_json_file(summary_path, build_report_summary(packet, results, report_path))
    if state_path.exists():
        state = load_run_state(state_path)
        state["overall_status"] = "aggregated"
        state["report_path"] = str(report_path)
        state["summary_json_path"] = str(summary_path)
        save_run_state(state_path, state)
    print(f"已生成报告: {report_path}")
    print(f"已生成摘要JSON: {summary_path}")
    print(f"已汇总 OR 数: {len(results)}")


def main(argv: Sequence[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description="Shard-only requirement evaluation helper.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    build_parser = subparsers.add_parser("build-packet", help="Build packet.json from Excel/JSON input.")
    build_parser.add_argument("--input", required=True, help="Path to the input Excel/JSON file.")
    build_parser.add_argument("--output", help="Path to write packet.json.")
    build_parser.set_defaults(func=command_build_packet)

    split_parser = subparsers.add_parser("split-shards", help="Split an existing packet.json into shard JSON and prompt files.")
    split_parser.add_argument("--packet", required=True, help="Path to packet.json.")
    split_parser.add_argument("--out-dir", required=True, help="Directory for shards/ and prompts/.")
    split_parser.add_argument("--shard-size", type=int, default=DEFAULT_SHARD_SIZE, help="OR units per shard.")
    split_parser.set_defaults(func=command_split_shards)

    prepare_parser = subparsers.add_parser("prepare", help="Build packet, split shards, and generate prompts.")
    prepare_parser.add_argument("--input", required=True, help="Path to the input Excel/JSON file.")
    prepare_parser.add_argument("--out-dir", required=True, help="Directory for packet/shards/prompts/results/repairs.")
    prepare_parser.add_argument("--shard-size", type=int, default=DEFAULT_SHARD_SIZE, help="OR units per shard.")
    prepare_parser.add_argument("--state-json", help="Optional path to write the orchestration state JSON.")
    prepare_parser.set_defaults(func=command_prepare)

    validate_parser = subparsers.add_parser("validate-result", help="Validate raw model TSV output for one shard.")
    validate_parser.add_argument("--shard", required=True, help="Path to shard JSON.")
    validate_parser.add_argument("--raw-output", required=True, help="Path to raw model TSV output.")
    validate_parser.add_argument("--output", required=True, help="Path to write validation JSON.")
    validate_parser.add_argument("--repair-prompt", help="Optional path to write repair prompt when output is repairable.")
    validate_parser.add_argument("--state-json", help="Optional path to the orchestration state JSON to update.")
    validate_parser.add_argument(
        "--attempt-kind",
        choices=("raw", "repair", "rerun"),
        default="raw",
        help="Kind of model attempt being validated.",
    )
    validate_parser.set_defaults(func=command_validate_result)

    aggregate_parser = subparsers.add_parser("aggregate", help="Aggregate validated shard result JSON files into a Markdown report.")
    aggregate_parser.add_argument("--packet", required=True, help="Path to packet.json.")
    aggregate_parser.add_argument("--results-dir", required=True, help="Directory containing *.valid.json files.")
    aggregate_parser.add_argument("--report", required=True, help="Path to write final Markdown report.")
    aggregate_parser.add_argument("--summary-json", help="Optional path to write machine-readable summary JSON.")
    aggregate_parser.add_argument("--state-json", help="Optional path to the orchestration state JSON to enforce before aggregation.")
    aggregate_parser.set_defaults(func=command_aggregate)

    args = parser.parse_args(argv)
    args.func(args)


if __name__ == "__main__":
    main()
