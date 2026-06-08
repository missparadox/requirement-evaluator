import importlib.util
import json
import sys
import tempfile
import unittest
from pathlib import Path


MODULE_PATH = Path(__file__).resolve().parents[1] / "scripts" / "evaluate_requirements.py"


def load_module():
    spec = importlib.util.spec_from_file_location("requirements_eval_test", MODULE_PATH)
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


class RequirementsEvaluatorPacketTests(unittest.TestCase):
    def setUp(self):
        self.module = load_module()

    def test_build_dimensions_returns_default_rubric(self):
        dimensions = self.module.build_dimensions()
        by_key = {item["key"]: item for item in dimensions}
        self.assertEqual(by_key["or_user_language"]["weight"], 12)
        self.assertEqual(by_key["dr_security"]["weight"], 5)
        self.assertEqual(by_key["dr_technical"]["weight"], 10)
        self.assertEqual(by_key["dr_ambiguity"]["weight"], 8)
        self.assertEqual(by_key["dr_exception"]["weight"], 7)
        self.assertEqual(by_key["cross_scope"]["weight"], 7)
        self.assertEqual(by_key["cross_dependencies"]["weight"], 6)
        self.assertEqual(by_key["cross_traceability"]["weight"], 7)
        self.assertEqual(sum(item["weight"] for item in dimensions if item["key"].startswith("or_")), 40)
        self.assertEqual(sum(item["weight"] for item in dimensions if item["key"].startswith("dr_")), 40)
        self.assertEqual(sum(item["weight"] for item in dimensions if item["key"].startswith("cross_")), 20)
        self.assertEqual(by_key["or_scenario"]["name"], "OR-应用场景")
        self.assertNotIn("dr_performance", by_key)
        self.assertNotIn("dr_hardware", by_key)
        self.assertNotIn("cross_exceptions", by_key)

    def test_build_review_packet_keeps_rows_and_core_fields(self):
        record_1 = self.module.RowRecord(
            index=1,
            grouped={
                "OR需求编号": ["DOR-1"],
                "OR需求名称*": ["网络检测与诊断"],
                "OR需求描述*": ["提供网络检测与维护功能。"],
                "分类类型": ["功能"],
                "需求来源": ["客户定制"],
                "国家/区域": ["中国"],
                "DR需求编号": ["DDR-1"],
                "DR需求名称*": ["Ping检测"],
                "DR需求描述*": ["支持 Ping 检测。"],
                "参数规格": ["次数 1-10"],
                "系统测试要点": ["校验 Ping 输入"],
                "DS需求名称*": ["Ping规格"],
                "规格分类": ["接口规格"],
                "所属子系统": ["网络管理"],
            },
        )
        record_2 = self.module.RowRecord(
            index=2,
            grouped={
                "OR需求编号": ["DOR-1"],
                "OR需求名称*": ["网络检测与诊断"],
                "OR需求描述*": ["提供网络检测与维护功能。"],
                "分类类型": ["功能"],
                "需求来源": ["客户定制"],
                "国家/区域": ["中国"],
                "DR需求编号": ["DDR-2"],
                "DR需求名称*": ["Telnet检测"],
                "DR需求描述*": ["支持 Telnet 检测。"],
                "参数规格": ["端口 1-65535"],
                "系统测试要点": ["校验 Telnet 输入"],
                "DS需求名称*": ["Telnet规格"],
                "规格分类": ["接口规格"],
                "所属子系统": ["网络管理"],
            },
        )

        packet = self.module.build_review_packet(
            input_path=Path("requirements.json"),
            dimensions=self.module.DEFAULT_DIMENSIONS,
            records=[record_1, record_2],
            source_info={"input_format": "json"},
        )

        self.assertEqual(packet["item_count"], 1)
        self.assertEqual(packet["or_count"], 1)
        self.assertEqual(packet["dr_count"], 2)
        self.assertEqual(packet["excluded_or_count"], 0)
        self.assertEqual(packet["all_category_counts"][0]["category"], "功能")
        self.assertTrue(packet["all_category_counts"][0]["included_in_evaluation"])
        self.assertEqual(packet["score_structure"]["or_total_weight"], 40)
        self.assertEqual(packet["score_structure"]["dr_total_weight"], 40)
        self.assertEqual(packet["score_structure"]["cross_total_weight"], 20)
        self.assertEqual(packet["source_info"]["input_format"], "json")
        self.assertEqual(packet["groups"][0]["id"], "DOR-1")
        self.assertEqual(packet["groups"][0]["category"], "功能")
        self.assertEqual(packet["groups"][0]["category_field"], "分类类型")
        self.assertIn("raw_fields", packet["groups"][0])
        self.assertIn("or_dimension_view", packet["groups"][0])
        self.assertIn("cross_dimension_view", packet["groups"][0])
        self.assertEqual(packet["groups"][0]["or_core_fields"]["requirement_source"], "客户定制")
        self.assertEqual(packet["groups"][0]["or_core_fields"]["region"], "中国")
        self.assertEqual(packet["groups"][0]["dr_count"], 2)
        self.assertEqual(packet["groups"][0]["review_skeleton"]["or_total_score"]["max_score"], 100)
        self.assertEqual(packet["groups"][0]["review_skeleton"]["or_part"]["max_score"], 40)
        self.assertEqual(packet["groups"][0]["review_skeleton"]["dr_average"]["max_score"], 40)
        self.assertEqual(packet["groups"][0]["review_skeleton"]["decomposition_quality"]["max_score"], 20)
        self.assertEqual(len(packet["groups"][0]["review_skeleton"]["dr_parts"]), 2)
        self.assertEqual(packet["groups"][0]["dr_items"][0]["core_fields"]["dr_desc"], "支持 Ping 检测。")
        self.assertEqual(packet["groups"][0]["dr_items"][0]["core_fields"]["spec_type"], "接口规格")
        self.assertEqual(packet["groups"][0]["dr_items"][0]["core_fields"]["subsystem"], "网络管理")
        self.assertIn("dr_testability", packet["groups"][0]["dr_items"][0]["dimension_view"])
        self.assertEqual(packet["groups"][0]["or_dimension_view"]["or_constraints"]["evidence_fields"]["国家/区域"], ["中国"])
        self.assertEqual(packet["groups"][0]["dr_items"][0]["dimension_view"]["dr_technical"]["evidence_fields"]["规格分类"], ["接口规格"])
        self.assertEqual(packet["groups"][0]["cross_dimension_view"]["cross_dependencies"]["evidence_fields"]["所属子系统"], ["网络管理"])

    def test_dimension_view_is_based_on_rubric_relevance_not_on_non_empty_frequency(self):
        record = self.module.RowRecord(
            index=1,
            grouped={
                "OR需求编号": ["DOR-2"],
                "OR需求名称*": ["设备接入"],
                "OR需求描述*": ["支持设备接入。"],
                "分类类型": ["功能"],
                "DR需求描述*": ["设备接入需校验参数格式。"],
                "参数规格": ["字段长度 1-64"],
            },
        )

        packet = self.module.build_review_packet(
            input_path=Path("sample.json"),
            dimensions=self.module.DEFAULT_DIMENSIONS,
            records=[record],
            source_info={"input_format": "json"},
        )
        dim_view = packet["groups"][0]["dr_items"][0]["dimension_view"]

        self.assertIn("dr_testability", dim_view)
        self.assertIn("系统测试要点", dim_view["dr_testability"]["mapped_fields"])
        self.assertIn("系统测试要点", dim_view["dr_testability"]["missing_fields"])
        self.assertEqual(dim_view["dr_testability"]["evidence_fields"]["参数规格"], ["字段长度 1-64"])

    def test_build_review_packet_only_evaluates_functional_requirements(self):
        functional = self.module.RowRecord(
            index=1,
            grouped={
                "OR需求编号": ["DOR-1"],
                "OR需求名称*": ["功能需求"],
                "OR需求描述*": ["提供设备接入功能。"],
                "分类类型": ["功能"],
                "DR需求编号": ["DDR-1"],
                "DR需求描述*": ["校验设备接入参数。"],
            },
        )
        non_functional = self.module.RowRecord(
            index=2,
            grouped={
                "OR需求编号": ["DOR-2"],
                "OR需求名称*": ["性能需求"],
                "OR需求描述*": ["响应时间小于 1 秒。"],
                "分类类型": ["性能"],
                "DR需求编号": ["DDR-2"],
                "DR需求描述*": ["接口响应时间小于 1 秒。"],
            },
        )

        packet = self.module.build_review_packet(
            input_path=Path("sample.json"),
            dimensions=self.module.DEFAULT_DIMENSIONS,
            records=[functional, non_functional],
            source_info={"input_format": "json"},
        )

        self.assertEqual(packet["item_count"], 1)
        self.assertEqual(packet["or_count"], 1)
        self.assertEqual(packet["excluded_or_count"], 1)
        self.assertEqual(packet["groups"][0]["id"], "DOR-1")
        by_category = {item["category"]: item for item in packet["all_category_counts"]}
        self.assertTrue(by_category["功能"]["included_in_evaluation"])
        self.assertFalse(by_category["性能"]["included_in_evaluation"])
        self.assertEqual(by_category["性能"]["exclusion_reason"], "分类类型不是功能")

    def test_cli_build_packet_writes_json_packet(self):
        rows = [
            {
                "OR需求编号": "DOR-1",
                "OR需求名称*": "网络检测与诊断",
                "OR需求描述*": "提供网络检测与维护功能。",
                "分类类型": "功能",
                "DR需求编号": "DDR-1",
                "DR需求名称*": "Ping检测",
                "DR需求描述*": "支持 Ping 检测。",
            },
            {
                "OR需求编号": "DOR-1",
                "OR需求名称*": "网络检测与诊断",
                "OR需求描述*": "提供网络检测与维护功能。",
                "分类类型": "功能",
                "DR需求编号": "DDR-2",
                "DR需求名称*": "Telnet检测",
                "DR需求描述*": "支持 Telnet 检测。",
            },
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            input_path = tmp / "requirements.json"
            output_path = tmp / "packet.json"
            input_path.write_text(json.dumps(rows, ensure_ascii=False), encoding="utf-8")

            self.module.main(
                [
                    "build-packet",
                    "--input",
                    str(input_path),
                    "--output",
                    str(output_path),
                ]
            )

            packet = json.loads(output_path.read_text(encoding="utf-8"))

        self.assertEqual(packet["item_count"], 1)
        self.assertEqual(packet["or_count"], 1)
        self.assertEqual(packet["dr_count"], 2)
        self.assertEqual(packet["source_info"]["input_format"], "json")
        self.assertEqual(packet["groups"][0]["name"], "网络检测与诊断")
        self.assertIn("review_skeleton", packet["groups"][0])

    def test_cli_build_packet_default_output_uses_input_file_stem(self):
        rows = [
            {
                "OR需求编号": "DOR-1",
                "OR需求名称*": "网络检测与诊断",
                "OR需求描述*": "提供网络检测与维护功能。",
                "分类类型": "功能",
                "DR需求编号": "DDR-1",
                "DR需求名称*": "Ping检测",
                "DR需求描述*": "支持 Ping 检测。",
            }
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            input_path = tmp / "requirements.json"
            output_path = tmp / "requirements.json"
            input_path.write_text(json.dumps(rows, ensure_ascii=False), encoding="utf-8")

            self.module.main(["build-packet", "--input", str(input_path)])

            packet = json.loads(output_path.read_text(encoding="utf-8"))

        self.assertEqual(packet["or_count"], 1)

    def test_prepare_generates_packet_shards_prompts_and_result_dirs(self):
        rows = [
            {
                "OR需求编号": "DOR-1",
                "OR需求名称*": "网络检测",
                "OR需求描述*": "提供网络检测功能。",
                "分类类型": "功能",
                "DR需求编号": "DDR-1",
                "DR需求描述*": "支持 Ping 检测。",
            },
            {
                "OR需求编号": "DOR-2",
                "OR需求名称*": "诊断导出",
                "OR需求描述*": "提供诊断导出功能。",
                "分类类型": "功能",
                "DR需求编号": "DDR-2",
                "DR需求描述*": "支持导出诊断日志。",
            },
            {
                "OR需求编号": "DOR-3",
                "OR需求名称*": "远程维护",
                "OR需求描述*": "提供远程维护功能。",
                "分类类型": "功能",
                "DR需求编号": "DDR-3",
                "DR需求描述*": "支持远程维护操作。",
            },
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            input_path = tmp / "requirements.json"
            out_dir = tmp / "artifacts"
            input_path.write_text(json.dumps(rows, ensure_ascii=False), encoding="utf-8")

            self.module.main(["prepare", "--input", str(input_path), "--out-dir", str(out_dir), "--shard-size", "2"])

            packet = json.loads((out_dir / "packet.json").read_text(encoding="utf-8"))
            shard_1 = json.loads((out_dir / "shards" / "shard_0001.json").read_text(encoding="utf-8"))
            shard_2 = json.loads((out_dir / "shards" / "shard_0002.json").read_text(encoding="utf-8"))
            prompt = (out_dir / "prompts" / "shard_0001.prompt.txt").read_text(encoding="utf-8")
            results_dir_exists = (out_dir / "results").exists()
            repairs_dir_exists = (out_dir / "repairs").exists()

        self.assertEqual(packet["or_count"], 3)
        self.assertEqual(shard_1["expected_or_ids"], ["DOR-1", "DOR-2"])
        self.assertEqual(shard_2["expected_or_ids"], ["DOR-3"])
        self.assertIn("[SCORING_GUIDE]", prompt)
        self.assertIn("[OUTPUT_TSV_SCHEMA]", prompt)
        self.assertIn("[SHARD_JSON]", prompt)
        self.assertTrue(results_dir_exists)
        self.assertTrue(repairs_dir_exists)

    def test_validate_result_accepts_valid_tsv(self):
        shard = {
            "shard_id": "shard_0001",
            "expected_or_ids": ["DOR-1", "DOR-2"],
        }
        raw_tsv = (
            "or_id\tor_name\ttotal_score\tor_score\tdr_average_score\ttraceability_score\tgrade\tweak_dimensions\tred_flags\tmissing_items\trevision_actions\tevidence_summary\n"
            "DOR-1\t网络检测\t72\t28\t31\t13\tfair\tDR-异常描述\t缺少异常路径\t验收条件\t补充异常场景\t基本行为明确，但异常和验收不足。\n"
            "DOR-2\t诊断导出\t81\t32\t34\t15\tgood\tOR-约束和限制\t无\t边界条件\t补充导出边界\t导出行为较清楚，但边界不完整。\n"
        )

        validation = self.module.validate_tsv_output(shard, raw_tsv)

        self.assertEqual(validation["status"], "valid")
        self.assertEqual(len(validation["results"]), 2)
        self.assertEqual(validation["results"][0]["total_score"], 72)
        self.assertEqual(validation["results"][0]["weak_dimensions"], ["DR-异常描述"])

    def test_validate_result_marks_markdown_table_as_repairable(self):
        shard = {
            "shard_id": "shard_0001",
            "expected_or_ids": ["DOR-1"],
        }
        raw_output = (
            "| or_id | or_name | total_score | or_score | dr_average_score | traceability_score | grade |\n"
            "| --- | --- | ---: | ---: | ---: | ---: | --- |\n"
            "| DOR-1 | 网络检测 | 72 | 28 | 31 | 13 | fair |\n"
        )

        validation = self.module.validate_tsv_output(shard, raw_output)

        self.assertEqual(validation["status"], "repairable")

    def test_validate_result_requires_rerun_when_or_id_is_missing(self):
        shard = {
            "shard_id": "shard_0001",
            "expected_or_ids": ["DOR-1"],
        }
        raw_output = "这里只是泛泛总结，没有结构化评分。"

        validation = self.module.validate_tsv_output(shard, raw_output)

        self.assertEqual(validation["status"], "rerun_required")

    def test_aggregate_renders_markdown_report_from_valid_results(self):
        packet = {
            "input_path": "requirements.json",
            "source_info": {"input_format": "json"},
            "or_count": 2,
            "dr_count": 2,
            "all_category_counts": [
                {
                    "category": "功能",
                    "count": 2,
                    "percentage": 100,
                    "included_in_evaluation": True,
                    "exclusion_reason": "",
                }
            ],
            "groups": [
                {"id": "DOR-1", "name": "网络检测"},
                {"id": "DOR-2", "name": "诊断导出"},
            ],
        }
        results = [
            {
                "or_id": "DOR-1",
                "or_name": "网络检测",
                "total_score": 72,
                "or_score": 28,
                "dr_average_score": 31,
                "traceability_score": 13,
                "grade": "fair",
                "weak_dimensions": ["DR-异常描述"],
                "red_flags": ["缺少异常路径"],
                "missing_items": ["验收条件"],
                "revision_actions": ["补充异常场景"],
                "evidence_summary": "基本行为明确，但异常和验收不足。",
            }
        ]

        report = self.module.render_aggregate_report(packet, results)

        self.assertIn("# 需求评估报告", report)
        self.assertIn("- 成功评估 OR 数: 1", report)
        self.assertIn("| DOR-1 | 网络检测 | 72 | 28 | 31 | 13 | fair |", report)
        self.assertIn("| DOR-2 | 诊断导出 | 未评估 |", report)
        self.assertIn("## 9. 未评估或失败项", report)

    def test_read_excel_records_active_sheet_name_in_source_info(self):
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl is not installed in the current test environment")

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            input_path = tmp / "requirements.xlsx"

            workbook = openpyxl.Workbook()
            default_sheet = workbook.active
            default_sheet.title = "Sheet1"
            default_sheet["A1"] = "OR需求编号"
            default_sheet["B1"] = "OR需求名称*"
            default_sheet["C1"] = "OR需求描述*"
            default_sheet["D1"] = "分类类型"
            default_sheet["E1"] = "DR需求编号"
            default_sheet["F1"] = "DR需求名称*"
            default_sheet["G1"] = "DR需求描述*"

            default_sheet["A2"] = "DOR-1"
            default_sheet["B2"] = "网络检测与诊断"
            default_sheet["C2"] = "提供网络检测与维护功能。"
            default_sheet["D2"] = "功能"
            default_sheet["E2"] = "DDR-1"
            default_sheet["F2"] = "Ping检测"
            default_sheet["G2"] = "支持 Ping 检测。"
            default_sheet["E3"] = "DDR-2"
            default_sheet["F3"] = "Telnet检测"
            default_sheet["G3"] = "支持 Telnet 检测。"
            default_sheet.merge_cells("A2:A3")
            default_sheet.merge_cells("B2:B3")
            default_sheet.merge_cells("C2:C3")
            default_sheet.merge_cells("D2:D3")
            workbook.create_sheet("OtherSheet")
            workbook.save(input_path)

            result = self.module.read_excel(input_path)

        self.assertEqual(result.source_info["sheet_name"], "Sheet1")
        self.assertEqual(result.source_info["input_format"], "xlsx")
        self.assertEqual(len(result.records), 2)
        self.assertEqual(result.records[1].first("OR需求编号"), "DOR-1")
        self.assertEqual(result.records[1].first("OR需求名称*"), "网络检测与诊断")


    def test_missing_excel_dependency_message_contains_install_command(self):
        original_find_spec = self.module.importlib.util.find_spec

        def fake_find_spec(name):
            if name == "openpyxl":
                return None
            return original_find_spec(name)

        self.module.importlib.util.find_spec = fake_find_spec
        try:
            with self.assertRaises(SystemExit) as ctx:
                self.module.ensure_runtime_dependencies(Path("input.xlsx"))
        finally:
            self.module.importlib.util.find_spec = original_find_spec

        message = str(ctx.exception)
        self.assertIn("openpyxl", message)
        self.assertIn("python3 -m pip install openpyxl", message)


if __name__ == "__main__":
    unittest.main()
